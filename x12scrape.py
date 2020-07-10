#!/usr/bin/python3

import requests
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import Workbook
from urllib.parse import urlparse

def get_title_from_url(url):
    u = urlparse(url)
    path = u.path
    split_path = path.split('/')
    return split_path[-1:][0].replace('-', ' ').title()

def scrape_codes(url):
    page = requests.get(url)
    soup = BeautifulSoup(page.content, 'html.parser')
    codelist = soup.find(id="codelist")

    if codelist is None:
        return None

    prod_set_current_codes = codelist.select(".prod_set.current")
    codes = {}

    for c in prod_set_current_codes:
        cas = c.select("td.code")
        desc = c.select("td.description")
        codes.update({f"{cas[0].get_text()}": f"{desc[0].get_text()}"})

    title = get_title_from_url(url)
    title = title[:30]

    return [codes, title]

def create_codefile(urls):
    wb = Workbook()

    ws = wb.active

    for u in urls:
        i = 1
        scraped_codes = scrape_codes(u)
        if scraped_codes is None:
            continue
        codes = scraped_codes[0]
        title = scraped_codes[1]
        for c in codes:
            ws.cell(i, 1, c)
            ws.cell(i, 2, codes[c])
            i += 1
        if title:
            ws.title = title
        ws = wb.create_sheet()

    wb.save("x12codes.xlsx")

def get_urls():
    page = requests.get("https://nex12.org/index.php/codes")
    soup = BeautifulSoup(page.content, 'html.parser')
    content = soup.find(id="content")
    link_table = content.select(".item-page table")

    a_tags = link_table[0].find_all("a")
    return [a.get("href") for a in a_tags]


urls = get_urls()
create_codefile(urls)
